"""
Module for generating Excel reports with cryptocurrency analysis.
Creates formatted spreadsheets with statistical data for different time periods.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import Dict, List

class ExcelReporter:
    """Generates Excel reports from statistical analysis data."""
    
    NUMBER_FORMAT_DECIMAL = '#,##0.00'
    PERIODS = ["12_months", "6_months", "3_months", "1_month"]
    PERIOD_DISPLAY = {
        "12_months": "12M",
        "6_months": "6M",
        "3_months": "3M",
        "1_month": "1M",
    }
    
    def __init__(self, filename: str = "reports/AnaliseCrypto.xlsx"):
        """
        Initialize the Excel reporter.
        
        Args:
            filename: Output Excel file path
        """
        self.filename = filename
        self.workbook = openpyxl.Workbook()
        self.workbook.remove(self.workbook.active)  # Remove default sheet
    
    def _setup_column_widths(self, ws):
        """Set up column widths for the summary sheet."""
        ws.column_dimensions['A'].width = 3  # Favorite column
        ws.column_dimensions['B'].width = 6.57  # Symbol column (46 pixels)
        ws.column_dimensions['C'].width = 9  # Latest Quote column
        ws.column_dimensions['D'].width = 9  # Second Latest Quote column
        ws.column_dimensions['E'].width = 5  # Period column
        
        # Statistics columns F to J (5 columns) - width 8.5
        for i in range(5):
            col_letter = get_column_letter(6 + i)
            ws.column_dimensions[col_letter].width = 8.5
        
        # Percentage columns K to N (4 columns) - 55 pixels = 7.86 units
        for i in range(4):
            col_letter = get_column_letter(11 + i)
            ws.column_dimensions[col_letter].width = 7.86
        
        # Statistics columns O to Q (3 columns) - width 8.5
        for i in range(3):
            col_letter = get_column_letter(15 + i)
            ws.column_dimensions[col_letter].width = 8.5
        
        # Percentage columns R to U (4 columns) - 55 pixels = 7.86 units
        for i in range(4):
            col_letter = get_column_letter(18 + i)
            ws.column_dimensions[col_letter].width = 7.86
        
        # Volatility columns V to AA (6 columns) - adjust widths
        ws.column_dimensions['V'].width = 7  # Vol% (volatility)
        for i in range(5):  # W to AA (±5%, ±10%, ±15%, ±20%, Score/M)
            col_letter = get_column_letter(23 + i)
            ws.column_dimensions[col_letter].width = 5.29
    
    def _create_title_rows(self, ws):
        """Create title and date rows."""
        ws['A1'] = "Análise de Criptomoedas em EUR"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:AA1')
        ws.row_dimensions[1].height = 25
        
        ws['A2'] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        ws['A2'].font = Font(size=10, italic=True)
        ws.row_dimensions[2].height = 18
    
    def _create_headers(self, ws, row: int, header_fill, header_font, border):
        """Create column headers for the new row-based layout."""
        headers = ["Fav", "Symbol", "Last", "2nd Last", "Period",
                  "MIN", "MAX", "AVG", "STD", "AVG-STD",
                  "Last-AVG%", "Last-A-S%", "2nd-AVG%", "2nd-A-S%",
                  "MEDIAN", "MAD", "MED-MAD",
                  "Last-MED%", "Last-M-M%", "2nd-MED%", "2nd-M-M%",
                  "Vol%", "±5%", "±10%", "±15%", "±20%", "Score/M"]
        
        for i, header in enumerate(headers):
            col_letter = get_column_letter(i + 1)
            cell = ws[f'{col_letter}{row}']
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
    
    def _write_symbol_period_row(self, ws, row: int, symbol: str, period: str, report: Dict, 
                                  favorites: List[str], border):
        """Write data for a single cryptocurrency symbol and period combination."""
        period_data = report.get("periods", {}).get(period, {})
        
        # Favorite marker with class (in all rows)
        favorite_marker = ""
        favorite_color = None
        
        if isinstance(favorites, dict):
            # New format: dict with classes as keys
            if symbol in favorites.get('A', []):
                favorite_marker = "A"
                favorite_color = "FFD700"  # Gold for A
            elif symbol in favorites.get('B', []):
                favorite_marker = "B"
                favorite_color = "FFA500"  # Orange for B
            elif symbol in favorites.get('C', []):
                favorite_marker = "C"
                favorite_color = "87CEEB"  # Light blue for C
        else:
            # Legacy format: list of symbols (treat as class A)
            if symbol in favorites:
                favorite_marker = "X"
                favorite_color = "FFD700"
        
        ws[f'A{row}'] = favorite_marker
        ws[f'A{row}'].font = Font(bold=True, size=12)
        if favorite_color:
            ws[f'A{row}'].fill = PatternFill(start_color=favorite_color, end_color=favorite_color, fill_type="solid")
        ws[f'A{row}'].border = border
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        
        # Symbol (in all rows)
        ws[f'B{row}'] = symbol
        ws[f'B{row}'].font = Font(bold=True)
        ws[f'B{row}'].border = border
        
        # Latest quote (column C)
        latest_quote = period_data.get("latest_quote")
        ws[f'C{row}'] = latest_quote
        ws[f'C{row}'].number_format = self.NUMBER_FORMAT_DECIMAL
        ws[f'C{row}'].border = border
        ws[f'C{row}'].alignment = Alignment(horizontal='right')
        ws[f'C{row}'].font = Font(bold=True, size=9)
        
        # Second latest quote (column D)
        second_latest_quote = period_data.get("second_latest_quote")
        ws[f'D{row}'] = second_latest_quote
        ws[f'D{row}'].number_format = self.NUMBER_FORMAT_DECIMAL
        ws[f'D{row}'].border = border
        ws[f'D{row}'].alignment = Alignment(horizontal='right')
        ws[f'D{row}'].font = Font(bold=True, size=9)
        
        # Period (column E)
        ws[f'E{row}'] = self.PERIOD_DISPLAY[period]
        ws[f'E{row}'].font = Font(bold=True, size=9)
        ws[f'E{row}'].border = border
        ws[f'E{row}'].alignment = Alignment(horizontal='center')
    def _write_period_stats(self, ws, row: int, period_data: Dict, border):
        """Write statistics for a specific period in the new row layout."""
        stats = period_data.get("stats", {})
        small_font = Font(size=9)
        
        # Minimum (column F)
        ws[f'F{row}'].value = stats.get("min")
        ws[f'F{row}'].number_format = self.NUMBER_FORMAT_DECIMAL
        ws[f'F{row}'].border = border
        ws[f'F{row}'].alignment = Alignment(horizontal='right')
        ws[f'F{row}'].font = small_font
        
        # Maximum (column G)
        ws[f'G{row}'].value = stats.get("max")
        ws[f'G{row}'].number_format = self.NUMBER_FORMAT_DECIMAL
        ws[f'G{row}'].border = border
        ws[f'G{row}'].alignment = Alignment(horizontal='right')
        ws[f'G{row}'].font = small_font
        
        # Mean (column H)
        ws[f'H{row}'].value = stats.get("mean")
        ws[f'H{row}'].number_format = self.NUMBER_FORMAT_DECIMAL
        ws[f'H{row}'].border = border
        ws[f'H{row}'].alignment = Alignment(horizontal='right')
        ws[f'H{row}'].font = small_font
        
        # Standard deviation (column I)
        ws[f'I{row}'].value = stats.get("std")
        ws[f'I{row}'].number_format = self.NUMBER_FORMAT_DECIMAL
        ws[f'I{row}'].border = border
        ws[f'I{row}'].alignment = Alignment(horizontal='right')
        ws[f'I{row}'].font = small_font
        
        # Mean - Std formula (column J)
        ws[f'J{row}'].value = f"=H{row}-I{row}"
        ws[f'J{row}'].number_format = self.NUMBER_FORMAT_DECIMAL
        ws[f'J{row}'].border = border
        ws[f'J{row}'].alignment = Alignment(horizontal='right')
        ws[f'J{row}'].font = small_font
        
        # Median (column O)
        ws[f'O{row}'].value = stats.get("median")
        ws[f'O{row}'].number_format = self.NUMBER_FORMAT_DECIMAL
        ws[f'O{row}'].border = border
        ws[f'O{row}'].alignment = Alignment(horizontal='right')
        ws[f'O{row}'].font = small_font
        
        # MAD - Median Absolute Deviation (column P)
        ws[f'P{row}'].value = stats.get("mad")
        ws[f'P{row}'].number_format = self.NUMBER_FORMAT_DECIMAL
        ws[f'P{row}'].border = border
        ws[f'P{row}'].alignment = Alignment(horizontal='right')
        ws[f'P{row}'].font = small_font
        
        # Median - MAD formula (column Q)
        ws[f'Q{row}'].value = f"=O{row}-P{row}"
        ws[f'Q{row}'].number_format = self.NUMBER_FORMAT_DECIMAL
        ws[f'Q{row}'].border = border
        ws[f'Q{row}'].alignment = Alignment(horizontal='right')
        ws[f'Q{row}'].font = small_font
        
        # Deviation formulas with conditional formatting
        self._write_deviation_formulas(ws, row, period_data, border)
    
    def _write_single_deviation_cell(self, ws, row: int, col: str, formula: str, 
                                     deviation_value, border):
        """Write a single deviation cell with formula and conditional formatting."""
        small_font = Font(size=9)
        ws[f'{col}{row}'].value = formula
        ws[f'{col}{row}'].number_format = '0.00%'
        ws[f'{col}{row}'].border = border
        ws[f'{col}{row}'].alignment = Alignment(horizontal='right')
        ws[f'{col}{row}'].font = small_font
        fill_color = "C6EFCE" if deviation_value and deviation_value >= 0 else "FFC7CE"
        ws[f'{col}{row}'].fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    
    def _write_deviation_formulas(self, ws, row: int, period_data: Dict, border):
        """Write deviation formulas with conditional formatting in the new row layout."""
        dev_mean_pct = period_data.get("latest_deviation_from_mean_pct")
        dev_mean_std_pct = period_data.get("latest_deviation_from_mean_minus_std_pct")
        second_dev_mean_pct = period_data.get("second_deviation_from_mean_pct")
        second_dev_mean_std_pct = period_data.get("second_deviation_from_mean_minus_std_pct")
        
        dev_median_pct = period_data.get("latest_deviation_from_median_pct")
        dev_median_mad_pct = period_data.get("latest_deviation_from_median_minus_mad_pct")
        second_dev_median_pct = period_data.get("second_deviation_from_median_pct")
        second_dev_median_mad_pct = period_data.get("second_deviation_from_median_minus_mad_pct")
        
        # Mean-based deviations
        self._write_single_deviation_cell(ws, row, 'K', f"=(C{row}-H{row})/H{row}", dev_mean_pct, border)
        self._write_single_deviation_cell(ws, row, 'L', f"=(C{row}-J{row})/J{row}", dev_mean_std_pct, border)
        self._write_single_deviation_cell(ws, row, 'M', f"=(D{row}-H{row})/H{row}", second_dev_mean_pct, border)
        self._write_single_deviation_cell(ws, row, 'N', f"=(D{row}-J{row})/J{row}", second_dev_mean_std_pct, border)
        
        # Median-based deviations
        self._write_single_deviation_cell(ws, row, 'R', f"=(C{row}-O{row})/O{row}", dev_median_pct, border)
        self._write_single_deviation_cell(ws, row, 'S', f"=(C{row}-Q{row})/Q{row}", dev_median_mad_pct, border)
        self._write_single_deviation_cell(ws, row, 'T', f"=(D{row}-O{row})/O{row}", second_dev_median_pct, border)
        self._write_single_deviation_cell(ws, row, 'U', f"=(D{row}-Q{row})/Q{row}", second_dev_median_mad_pct, border)
    
    def _write_volatility_stats(self, ws, row: int, volatility_data: Dict, period: str, border):
        """Write volatility statistics for each period row."""
        if not volatility_data:
            return
        
        small_font = Font(size=9)
        
        # Column V: Daily Volatility (annualized % from daily returns)
        daily_vol = volatility_data.get('daily_volatility')
        ws[f'V{row}'].value = daily_vol
        ws[f'V{row}'].number_format = '#,##0.00"%"'
        ws[f'V{row}'].border = border
        ws[f'V{row}'].alignment = Alignment(horizontal='center')
        ws[f'V{row}'].font = small_font
        
        # Column W: ±5% (format: positive:negative, e.g. "8:11")
        positive_5 = volatility_data.get('volatility_positive_5', 0)
        negative_5 = volatility_data.get('volatility_negative_5', 0)
        ws[f'W{row}'].value = f"{positive_5}:{negative_5}"
        ws[f'W{row}'].border = border
        ws[f'W{row}'].alignment = Alignment(horizontal='center')
        ws[f'W{row}'].font = small_font
        
        # Column X: ±10% (format: positive:negative, e.g. "8:11")
        positive_10 = volatility_data.get('volatility_positive_10', 0)
        negative_10 = volatility_data.get('volatility_negative_10', 0)
        ws[f'X{row}'].value = f"{positive_10}:{negative_10}"
        ws[f'X{row}'].border = border
        ws[f'X{row}'].alignment = Alignment(horizontal='center')
        ws[f'X{row}'].font = small_font
        
        # Column Y: ±15% (format: positive:negative, e.g. "8:11")
        positive_15 = volatility_data.get('volatility_positive_15', 0)
        negative_15 = volatility_data.get('volatility_negative_15', 0)
        ws[f'Y{row}'].value = f"{positive_15}:{negative_15}"
        ws[f'Y{row}'].border = border
        ws[f'Y{row}'].alignment = Alignment(horizontal='center')
        ws[f'Y{row}'].font = small_font
        
        # Column Z: ±20% (format: positive:negative, e.g. "8:11")
        positive_20 = volatility_data.get('volatility_positive_20', 0)
        negative_20 = volatility_data.get('volatility_negative_20', 0)
        ws[f'Z{row}'].value = f"{positive_20}:{negative_20}"
        ws[f'Z{row}'].border = border
        ws[f'Z{row}'].alignment = Alignment(horizontal='center')
        ws[f'Z{row}'].font = small_font
        
        # Column AA: Score/Mês (score por mês)
        period_months = {"12_months": 12, "6_months": 6, "3_months": 3, "1_month": 1}
        months = period_months.get(period, 1)
        score = volatility_data.get('volatility_score', 0)
        score_per_month = score / months if months > 0 else 0
        ws[f'AA{row}'].value = round(score_per_month, 1)
        ws[f'AA{row}'].border = border
        ws[f'AA{row}'].alignment = Alignment(horizontal='center')
        ws[f'Z{row}'].font = Font(bold=True, size=9)
        # Color code: higher score/month = more volatile
        if score_per_month > 25:
            ws[f'Z{row}'].fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        elif score_per_month > 15:
            ws[f'Z{row}'].fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    
    def create_summary_sheet(self, reports: Dict[str, Dict], market_caps: Dict[str, float] = None, favorites: List[str] = None):
        """
        Create a summary sheet with all cryptocurrencies and periods.
        Each cryptocurrency has 4 rows (one per period).
        
        Args:
            reports: Dictionary with analysis reports from StatisticalAnalyzer
            market_caps: Dictionary with market cap values for sorting (used for sorting)
            favorites: List of favorite cryptocurrency symbols (used for highlighting)
        """
        ws = self.workbook.create_sheet(title="Resumo")
        
        # Setup basic structure
        self._setup_column_widths(ws)
        self._create_title_rows(ws)
        
        # Style definitions
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=9)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Create column headers (row 4)
        self._create_headers(ws, 4, header_fill, header_font, border)
        ws.row_dimensions[4].height = 30  # Compact header row with top alignment
        
        # Sort symbols by market cap
        symbols = list(reports.keys())
        if market_caps:
            symbols = sorted(symbols, key=lambda s: market_caps.get(s, 0), reverse=True)
        else:
            symbols.sort()
        
        # Write data rows - 4 rows per symbol (one for each period)
        row = 5
        favorites = favorites or []
        for symbol in symbols:
            report = reports[symbol]
            
            # Write 4 rows for this symbol (one per period)
            for i, period in enumerate(self.PERIODS):
                # Write symbol, period, and quotes
                self._write_symbol_period_row(ws, row, symbol, period, report, favorites, border)
                
                # Write period statistics
                period_data = report.get("periods", {}).get(period, {})
                if period_data:
                    self._write_period_stats(ws, row, period_data, border)
                    
                    # Write volatility stats for this period
                    volatility_data = period_data.get('volatility', {})
                    self._write_volatility_stats(ws, row, volatility_data, period, border)
                
                row += 1
        
        # Add auto filter to the table
        ws.auto_filter.ref = f"A4:Z{row - 1}"
        
        # Freeze panes (freeze columns A-D and header row)
        ws.freeze_panes = ws['E5']
    
    def create_detail_sheet(self, symbol: str, report: Dict):
        """
        Create a detailed analysis sheet for a single cryptocurrency.
        
        Args:
            symbol: Cryptocurrency symbol
            report: Analysis report for the cryptocurrency
        """
        ws = self.workbook.create_sheet(symbol)
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title
        ws['A1'] = f"Análise Detalhada: {symbol}"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:D1')
        
        # Date range
        date_range = report.get("date_range", {})
        ws['A3'] = "Período de Dados:"
        ws['B3'] = f"{date_range.get('start', 'N/A')} até {date_range.get('end', 'N/A')}"
        ws['A4'] = "Total de Pontos de Dados:"
        ws['B4'] = report.get("data_points", 0)
        
        # Period analysis
        row = 6
        for period in self.PERIODS:
            period_display = self.PERIOD_DISPLAY[period]
            period_data = report.get("periods", {}).get(period, {})
            stats = period_data.get("stats", {})
            
            # Period header
            ws[f'A{row}'] = period_display
            ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
            ws[f'A{row}'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            ws.merge_cells(f'A{row}:B{row}')
            row += 1
            
            # Statistics
            metrics = [
                ("Mínimo", stats.get("min")),
                ("Máximo", stats.get("max")),
                ("Média", stats.get("mean")),
                ("Desvio Padrão", stats.get("std")),
                ("Média - Desvio Padrão", stats.get("mean_minus_std")),
                ("Última Cotação", period_data.get("latest_quote")),
                ("Desvio da Última Cotação à Média", period_data.get("latest_deviation_from_mean")),
                ("Desvio da Última Cotação à Média-Desvio", period_data.get("latest_deviation_from_mean_minus_std")),
                ("Total de Pontos", stats.get("count")),
            ]
            
            for metric_name, value in metrics:
                ws[f'A{row}'] = metric_name
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'A{row}'].border = border
                
                ws[f'B{row}'] = value
                ws[f'B{row}'].number_format = '0.00000000'
                ws[f'B{row}'].border = border
                
                row += 1
            
            row += 1  # Space between periods
        
        # Set column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 20
    
    def save(self):
        """Save the workbook to file."""
        import os
        os.makedirs(os.path.dirname(self.filename) or ".", exist_ok=True)
        self.workbook.save(self.filename)
        print(f"Excel report saved to: {self.filename}")
    
    def _write_volatility_detail_row(self, ws, row: int, favorite_class: str, symbol: str, 
                                     period: str, volatility_data: Dict, border) -> int:
        """Write a single volatility detail row with period information."""
        # Favorite marker with class
        favorite_marker = favorite_class if favorite_class else ""
        favorite_colors = {
            'A': "FFD700",  # Gold
            'B': "FFA500",  # Orange
            'C': "87CEEB"   # Light blue
        }
        
        ws.cell(row=row, column=1).value = favorite_marker
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        if favorite_class and favorite_class in favorite_colors:
            ws.cell(row=row, column=1).fill = PatternFill(
                start_color=favorite_colors[favorite_class], 
                end_color=favorite_colors[favorite_class], 
                fill_type="solid"
            )
        ws.cell(row=row, column=1).border = border
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
        
        # Symbol
        ws.cell(row=row, column=2).value = symbol
        ws.cell(row=row, column=2).font = Font(bold=True)
        ws.cell(row=row, column=2).border = border
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='left')
        
        # Period
        ws.cell(row=row, column=3).value = period
        ws.cell(row=row, column=3).border = border
        ws.cell(row=row, column=3).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=3).font = Font(bold=True)
        
        # Thresholds ordered by absolute variation with sum columns
        col = 4
        for threshold in [5, 10, 15, 20]:
            # Positive threshold
            pos_val = volatility_data.get(f'volatility_positive_{threshold}', 0)
            ws.cell(row=row, column=col).value = pos_val
            ws.cell(row=row, column=col).border = border
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
            col += 1
            
            # Negative threshold
            neg_val = volatility_data.get(f'volatility_negative_{threshold}', 0)
            ws.cell(row=row, column=col).value = neg_val
            ws.cell(row=row, column=col).border = border
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
            col += 1
            
            # Sum column (±threshold)
            sum_val = pos_val + neg_val
            ws.cell(row=row, column=col).value = sum_val
            ws.cell(row=row, column=col).border = border
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=col).font = Font(bold=True)
            ws.cell(row=row, column=col).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            col += 1
        
        # Score Weighted (com ponderação: 5*1.0, 10*1.5, 15*2.0, 20*2.5)
        score_weighted = volatility_data.get('volatility_score', 0)
        ws.cell(row=row, column=col).value = score_weighted
        ws.cell(row=row, column=col).border = border
        ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=col).font = Font(bold=True)
        if score_weighted > 100:
            ws.cell(row=row, column=col).fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        elif score_weighted > 50:
            ws.cell(row=row, column=col).fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        col += 1
        
        # Score/Mês (score dividido pelo número de meses)
        period_months = {"12M": 12, "6M": 6, "3M": 3, "1M": 1}
        months = period_months.get(period, 1)
        score_per_month = score_weighted / months if months > 0 else 0
        ws.cell(row=row, column=col).value = round(score_per_month, 1)
        ws.cell(row=row, column=col).border = border
        ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=col).font = Font(bold=True)
        
        return row
    
    def create_volatility_detail_sheet(self, reports: Dict[str, Dict], market_caps: Dict[str, float] = None, favorites: List[str] = None):
        """
        Create a detailed volatility analysis sheet organized by period.
        
        Args:
            reports: Dictionary with analysis reports including period-specific volatility
            market_caps: Dictionary with market cap values for sorting (same order as summary)
            favorites: List of favorite cryptocurrency symbols
        """
        ws = self.workbook.create_sheet(title="Volatility Detail")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        if favorites is None:
            favorites = []
        self._write_volatility_detail_title(ws)
        self._write_volatility_detail_headers(ws, border)
        self._set_volatility_detail_column_widths(ws)
        row = 4
        symbols = list(reports.keys())
        if market_caps:
            symbols = sorted(symbols, key=lambda s: market_caps.get(s, 0), reverse=True)
        else:
            symbols.sort()
        row = self._write_volatility_detail_data(ws, row, symbols, reports, favorites, border)
        if row > 4:
            ws.auto_filter.ref = f"A3:Q{row-1}"
        ws.freeze_panes = ws['A4']

    def _write_volatility_detail_title(self, ws):
        ws['A1'] = "Análise Detalhada de Volatilidade por Período"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:Q1')
        ws.row_dimensions[1].height = 25

    def _write_volatility_detail_headers(self, ws, border):
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        headers = ["Fav", "Symbol", "Period", "+5%", "-5%", "±5%", "+10%", "-10%", "±10%",
                  "+15%", "-15%", "±15%", "+20%", "-20%", "±20%", "Score", "Score/M"]
        for i, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=i)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='center')

    def _set_volatility_detail_column_widths(self, ws):
        ws.column_dimensions['A'].width = 4   # Fav
        ws.column_dimensions['B'].width = 9   # Symbol (63 pixels)
        ws.column_dimensions['C'].width = 6   # Period
        for col in ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O']:
            ws.column_dimensions[col].width = 7.43
        ws.column_dimensions['P'].width = 7.57   # Score
        ws.column_dimensions['Q'].width = 7.57   # Score/M

    def _write_volatility_detail_data(self, ws, row, symbols, reports, favorites, border):
        period_map = {
            "12M": "12_months",
            "6M": "6_months",
            "3M": "3_months",
            "1M": "1_month"
        }
        for symbol in symbols:
            row = self._write_symbol_volatility_rows(ws, row, symbol, reports, favorites, border, period_map)
        return row

    def _get_favorite_class(self, symbol, favorites):
        if isinstance(favorites, dict):
            for cls in ['A', 'B', 'C']:
                if symbol in favorites.get(cls, []):
                    return cls
        elif symbol in favorites:
            return 'A'  # Legacy format
        return None

    def _write_symbol_volatility_rows(self, ws, row, symbol, reports, favorites, border, period_map):
        periods = reports[symbol].get('periods', {})
        for period_label, period_key in period_map.items():
            if period_key in periods:
                period_data = periods[period_key]
                volatility_data = period_data.get('volatility', {})
                if volatility_data:
                    favorite_class = self._get_favorite_class(symbol, favorites)
                    row = self._write_volatility_detail_row(ws, row, favorite_class, symbol,
                                                            period_label, volatility_data, border)
                    row += 1
        return row
    
    def generate_report(self, reports: Dict[str, Dict], market_caps: Dict[str, float] = None, 
                       favorites: List[str] = None):
        """
        Generate complete Excel report.
        
        Args:
            reports: Dictionary with analysis reports from StatisticalAnalyzer
            market_caps: Dictionary with market cap values for sorting
            favorites: List of favorite cryptocurrency symbols
            volatility_results: (unused, deprecated)
        """
        # Create summary sheet
        self.create_summary_sheet(reports, market_caps, favorites)
        
        # Create volatility detail sheet with period-specific data
        self.create_volatility_detail_sheet(reports, market_caps, favorites)
        
        # Create detailed sheets for each cryptocurrency
        for symbol, report in sorted(reports.items()):
            if "error" not in report:
                self.create_detail_sheet(symbol, report)
        
        # Save the workbook
        self.save()
